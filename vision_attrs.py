# -*- coding: utf-8 -*-
"""
Vision v2 — 소재안 임베드 이미지에서 이미지 속성 추출.

파이프라인:
  1) extract_creative_images(bytes) : xlsx 임베드 이미지 → {(기획전번호,소재N): PIL.Image}
     (이미지는 '이미지' 컬럼(col 22)에 소재 행마다 앵커. 한 행에 여러 장이면 가장 큰 것=대표 소재컷.)
  2) cv_attrs(img)      : 무료 로컬(PIL) — 배경톤(웜/쿨/뉴트럴)·밝기구간. 키 불필요.
  3) gemini_attrs(img)  : Gemini 비전 1콜 → 인물수·상품표현·가격뱃지·타이포·레이아웃(JSON).
  4) analyze_images(...) : 위를 합쳐 (기획전번호,소재N) 키 DataFrame → 조인 테이블에 merge.

Gemini 키는 st.secrets["gemini_api_key"] 또는 env GEMINI_API_KEY. 무료 티어 쿼터를 위해
결과를 캐시(호출부에서 @st.cache_data)하고 배치로 돌린다.
"""
import io
import json
import re

_CODE = None  # lazy


# ══════════════════════════════════════════════════════════════════════
# 1. 임베드 이미지 추출 + 소재 매핑
# ══════════════════════════════════════════════════════════════════════
_PID_COL = 11   # K 기획전번호 (1-based)
# 이미지 앵커 컬럼(0-based) 우선순위: 최종소재안(31=완성 배너) > 이미지 피드백(30) > 이미지(22=구성컷).
# 최종운영안 파일엔 col31에 완성 배너(카피·로고·뱃지 포함)가 소재 행마다 들어있다.
_IMG_COL_PREF = [31, 30, 22]


def _sojae_row_map(ws, header_row=3):
    """엑셀행 → (기획전번호, 소재N). parse_sojae_bytes 와 동일한 그룹 내 순번 규칙."""
    out, cur, n = {}, None, 0
    for r in range(header_row + 1, ws.max_row + 1):
        pid = ws.cell(r, _PID_COL).value
        if pid in (None, ""):
            continue
        pid = str(pid).strip()
        if pid != cur:
            cur, n = pid, 0
        n += 1
        try:
            out[r] = (int(float(pid)), n)
        except ValueError:
            pass
    return out


def extract_creative_images(file_bytes, header_row=3, min_px=200):
    """xlsx bytes → {(기획전번호, 소재N): PIL.Image}. 최종 배너(col31) 우선, 행별 대표 1장.
    같은 행에 여러 배너(A/B)면 가장 큰 것. col31 없으면 col30→col22 폴백."""
    import openpyxl
    from PIL import Image
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)  # 이미지 로드 위해 non-readonly
    ws = wb[wb.sheetnames[0]]
    rowmap = _sojae_row_map(ws, header_row)
    # excel_row -> {col: (area, PIL)}
    per_row = {}
    for im in getattr(ws, "_images", []):
        try:
            col = im.anchor._from.col
            arow = im.anchor._from.row + 1
            data = im._data() if callable(getattr(im, "_data", None)) else im.ref
            pil = Image.open(io.BytesIO(data)).convert("RGB")
        except Exception:
            continue
        if pil.width < min_px and pil.height < min_px:
            continue
        d = per_row.setdefault(arow, {})
        area = pil.width * pil.height
        if col not in d or area > d[col][0]:
            d[col] = (area, pil)
    wb.close()
    out = {}
    for r, key in rowmap.items():
        cols = per_row.get(r)
        if not cols:
            continue
        for pref in _IMG_COL_PREF:                   # 최종소재안 우선
            if pref in cols:
                out[key] = cols[pref][1]
                break
        else:
            out[key] = max(cols.values(), key=lambda t: t[0])[1]
    return out


# ══════════════════════════════════════════════════════════════════════
# 2. 무료 로컬 CV (PIL) — 배경톤·밝기
# ══════════════════════════════════════════════════════════════════════
def cv_attrs(img):
    """배경톤(웜/쿨/뉴트럴)·밝기구간. 테두리 픽셀로 배경 추정."""
    from PIL import Image
    im = img.convert("RGB")
    im.thumbnail((160, 160))
    px = im.load()
    W, H = im.size
    # 전체 밝기
    tot = n = 0
    for y in range(0, H, 2):
        for x in range(0, W, 2):
            r, g, b = px[x, y]
            tot += 0.299 * r + 0.587 * g + 0.114 * b
            n += 1
    lum = tot / max(n, 1)
    bright = "어두움" if lum < 90 else ("밝음" if lum > 175 else "중간")
    # 배경(테두리 10%) 평균색
    m = max(4, int(min(W, H) * 0.1))
    rs = gs = bs = c = 0
    for y in range(H):
        for x in range(W):
            if x < m or x >= W - m or y < m or y >= H - m:
                r, g, b = px[x, y]
                rs += r; gs += g; bs += b; c += 1
    r, g, b = rs / c, gs / c, bs / c
    if r - b > 12:
        tone = "웜"
    elif b - r > 12:
        tone = "쿨"
    else:
        tone = "뉴트럴"
    return {"배경톤": tone, "밝기": bright, "배경밝기값": round(lum)}


# ══════════════════════════════════════════════════════════════════════
# 3. Gemini 비전 — 인물·상품표현·뱃지·타이포·레이아웃 (이미지당 1콜)
# ══════════════════════════════════════════════════════════════════════
_VISION_PROMPT = """이 이미지는 LF몰 디스플레이 광고 배너다. 아래 항목을 보고 JSON만 출력하라(설명 금지).
{
 "인물수": <정수, 사람이 몇 명 나오나. 없으면 0>,
 "상품표현": "<누끼|착용|연출|없음>",   // 누끼=배경제거 제품컷, 착용=모델 착용, 연출=연출/화보컷
 "가격뱃지": <true|false>,             // 가격/할인율이 뱃지·박스로 강조돼 있나
 "타이포강조": <true|false>,           // 큰 타이포(카피)가 이미지에 크게 박혀 있나
 "레이아웃": "<좌측형|중앙형|우측형>",  // 주 피사체/카피의 무게중심
 "인물성별": "<여성|남성|혼합|없음>"
}"""


def _gemini_key():
    import os
    try:
        import streamlit as st
        if "gemini_api_key" in st.secrets:
            return st.secrets["gemini_api_key"]
    except Exception:
        pass
    return os.getenv("GEMINI_API_KEY")


def gemini_attrs(img, api_key=None, model="gemini-3.6-flash", tries=4):
    """이미지 → 의미 속성 dict. 429(쿼터)면 지수백오프 재시도. 실패 시 {_error}."""
    import time
    key = api_key or _gemini_key()
    if not key:
        return {}
    try:
        from google import genai
        from google.genai import types
    except Exception as e:
        return {"_error": f"google-genai 미설치: {e}"}
    client = genai.Client(api_key=key)
    buf = io.BytesIO(); img.convert("RGB").save(buf, format="JPEG", quality=85)
    data = buf.getvalue()
    last = ""
    for attempt in range(tries):
        try:
            resp = client.models.generate_content(
                model=model,
                contents=[types.Part.from_bytes(data=data, mime_type="image/jpeg"), _VISION_PROMPT],
            )
            m = re.search(r"\{.*\}", (resp.text or "").strip(), re.S)
            return json.loads(m.group(0)) if m else {"_error": "JSON 파싱 실패"}
        except Exception as e:
            last = str(e)
            if "429" in last or "RESOURCE_EXHAUSTED" in last:
                time.sleep(min(2 ** attempt * 5, 40))   # 5,10,20,40s 백오프
                continue
            return {"_error": last[:150]}
    return {"_error": "429 재시도 초과: " + last[:100]}


# ══════════════════════════════════════════════════════════════════════
# 4. 배치 → DataFrame
# ══════════════════════════════════════════════════════════════════════
def build_vision_cache(svc, media="비즈보드", weeks=8, use_gemini=False, api_key=None,
                       sojae_root=None, progress=None):
    """SA로 최근 weeks 주차의 최종운영안(최신 파일) 배너를 추출→속성화→(기획전,소재N) 캐시 DataFrame.
    무겁다(파일당 non-readonly 로드 + 다수 이미지 + 선택적 Gemini) → 앱에선 캐시/버튼으로."""
    import pandas as pd
    import drive_fetch as D
    root = sojae_root or D.SOJAE_ROOT_ID
    weeks_f = sorted(D.list_children(svc, root, mime=D.FOLDER_MIME),
                     key=lambda x: x.get("createdTime", ""), reverse=True)[:weeks]
    frames = []
    for wk in weeks_f:
        sub = D.find_subfolder(svc, wk["id"], media)
        if not sub:
            continue
        xls = D._latest(D.list_children(svc, sub["id"], mime=D.XLSX_MIME), key="createdTime")
        if not xls:
            continue
        try:
            imgs = extract_creative_images(D.download_bytes(svc, xls["id"]))
        except Exception:
            continue
        if imgs:
            frames.append(analyze_images(imgs, use_gemini=use_gemini, api_key=api_key, progress=progress))
    if not frames:
        return pd.DataFrame()
    cat = pd.concat(frames, ignore_index=True).drop_duplicates(["기획전번호", "소재N"], keep="last")
    return cat.reset_index(drop=True)


def analyze_images(images, use_gemini=True, api_key=None, model="gemini-3.6-flash",
                   throttle=4.0, progress=None):
    """{(pid,N): PIL} → DataFrame[기획전번호,소재N,배경톤,밝기,인물수,상품표현,가격뱃지,타이포강조,레이아웃,인물성별].
    throttle: Gemini 콜 사이 최소 간격(초) — 무료 티어 RPM 제한 회피."""
    import pandas as pd, time
    rows = []
    items = list(images.items())
    for i, ((pid, N), img) in enumerate(items):
        rec = {"기획전번호": pid, "소재N": N}
        rec.update(cv_attrs(img))
        if use_gemini:
            g = gemini_attrs(img, api_key=api_key, model=model)
            rec.update({k: v for k, v in g.items() if not k.startswith("_")})
            if throttle and i < len(items) - 1:
                time.sleep(throttle)
        rows.append(rec)
        if progress:
            progress(i + 1, len(items))
    df = pd.DataFrame(rows)
    if not df.empty:
        df["기획전번호"] = df["기획전번호"].astype("Int64")
        df["소재N"] = df["소재N"].astype("Int64")
    return df
