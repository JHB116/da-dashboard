# -*- coding: utf-8 -*-
"""
DA 카피 분석·제안
─────────────────────────────────────────────────────────────
DA 디스플레이 광고 '소재(카피+이미지)'와 '실적'을 소재 단위로 조인해,
어떤 소구형·카피·이미지속성이 성과(CTR·ROAS·RPS)를 만드는지와 다음 제안을 도출한다.

── 데이터 모델 (실제 검증됨) ───────────────────────────────
  · 소재안(기획측): 주차별 Drive 엑셀(비즈보드/디스플레이/버즈빌 …).
      한 기획전(기획전번호) 안에서 소재가 등장 순서대로 1,2,3… 번호를 가진다.
      컬럼: 기획전번호·기획전명·브랜드·카테고리·성별·이미지유형·소구형(수동)·메인카피·서브카피 …
      ⚠️ 수동 '소구형'은 사람이 단 라벨이라 부정확 → 카피 원문/이미지 기준으로 재분류한다.
  · 실적(성과측): DA로우_RAW (25~26년 누적). 소재별 성과는
      구분_AF코드이름 = 'M카카오비즈보드_{기획전}(ADID[/앵커링])_{N}' 의 끝 _{N} 으로 식별.
      → (구분_기획전 번호, 소재순서 N) 이 소재안과의 조인 키.
      같은 소재 N 이 매체 변형(ADID / ADID앵커링)으로 여러 행이면 합산한다.
  · 조인: (기획전번호, 소재N)  →  소재의 {카피·소구형·이미지속성} ↔ {노출·클릭·거래액·ROAS}

── 설계 원칙 (발송 대시보드 계승) ──────────────────────────
  · 순수 함수(load_da_perf / parse_creative_order / tag_copy / appeal_perf …) = Streamlit 무관·테스트 가능.
  · 단일 하루는 소재별 전환이 희소(거래액 0 다수) → 분석은 '기간 누적' 위에서 해야 안정적.
"""
import io
import re
import numpy as np
import pandas as pd

# ══════════════════════════════════════════════════════════════════════
# 0-A. 소재안(주차별 Drive 엑셀) → 소재 단위 카피/속성
# ══════════════════════════════════════════════════════════════════════
# 소재안 시트 컬럼 인덱스(0-based, 헤더 3행). 실측 확인됨.
SOJAE_COLS = {
    "기획전번호": 10, "기획전명": 11, "상품코드": 12, "상품명": 13, "이미지유형": 14,
    "로고": 15, "브랜드": 16, "메인카피": 18, "서브카피": 20,
    "성별": 4, "카테고리": 8,
}


def parse_sojae_bytes(file_bytes, sheet=None, header_row=3):
    """소재안 엑셀 bytes → 소재 단위 DataFrame.
    한 기획전번호 그룹 안에서 등장 순서대로 소재N(1,2,3…)을 매긴다(실적 _N 과 대칭)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet] if sheet else wb[wb.sheetnames[0]]
    recs, cur_pid, n = [], None, 0
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        def cell(key):
            i = SOJAE_COLS[key]
            v = row[i] if i < len(row) else None
            return "" if v is None else str(v).replace("\n", " ").strip()
        pid = cell("기획전번호")
        if not pid:
            continue
        if pid != cur_pid:
            cur_pid, n = pid, 0
        n += 1
        recs.append({
            "기획전번호": pid, "소재N": n, "기획전명": cell("기획전명"),
            "브랜드": cell("브랜드"), "카테고리": cell("카테고리"), "성별": cell("성별"),
            "이미지유형": cell("이미지유형"), "로고": cell("로고"),
            "상품코드": cell("상품코드"), "상품명": cell("상품명"),
            "메인카피": cell("메인카피"), "서브카피": cell("서브카피"),
        })
    wb.close()
    df = pd.DataFrame(recs)
    if not df.empty:
        df["기획전번호"] = pd.to_numeric(df["기획전번호"], errors="coerce").astype("Int64")
        df["소재N"] = df["소재N"].astype("Int64")
    return df


# ══════════════════════════════════════════════════════════════════════
# 0-B. 실적(DA로우_RAW) → 소재별 성과
# ══════════════════════════════════════════════════════════════════════
# 매체 프리셋: 소재안 폴더(비즈보드/디스플레이/버즈빌)와 대응. AF코드이름 안의 매체 토큰으로 식별.
MEDIA_TOKENS = {"비즈보드": "비즈보드", "버즈빌": "버즈빌", "디스플레이": "디스플레이"}

PERF_NUM = ["지표_노출수", "지표_클릭수", "지표_광고비", "지표_순결제거래액",
            "지표_순결제고객수", "지표_UV(전체)", "지표_총결제거래액", "지표_순결제고객수(첫구매)"]

# 성과 주지표: 라벨 → (분자, 분모, 클수록 좋은가)
METRICS = {
    "ROAS":   ("지표_순결제거래액", "지표_광고비",   True),
    "CTR":    ("지표_클릭수",       "지표_노출수",   True),
    "RPS(노출당)": ("지표_순결제거래액", "지표_노출수", True),
    "CR(순)": ("지표_순결제고객수",  "지표_UV(전체)", True),
    "CPC(낮을수록↑)": ("지표_광고비", "지표_클릭수",  False),
}


def parse_creative_order(af_name):
    """구분_AF코드이름 끝의 _{N} → 소재순서(int) 또는 NA."""
    m = re.search(r"_(\d+)\s*$", str(af_name))
    return int(m.group(1)) if m else pd.NA


def load_da_perf(raw_df, media="비즈보드"):
    """DA로우_RAW → (기획전번호, 소재N) 단위 소재별 성과 집계.
    같은 소재의 매체변형(ADID/앵커링) 행은 합산한다."""
    df = raw_df.copy()
    name = df["구분_AF코드이름"].astype(str)
    if media and media in MEDIA_TOKENS:
        df = df[name.str.contains(MEDIA_TOKENS[media], na=False)].copy()
    df["소재N"] = df["구분_AF코드이름"].map(parse_creative_order).astype("Int64")
    df["기획전번호"] = pd.to_numeric(df["구분_기획전 번호"], errors="coerce").astype("Int64")
    for c in PERF_NUM:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)
    keys = ["기획전번호", "소재N"]
    nm = df.groupby(keys, dropna=False)["구분_브랜드/기획전"].first()
    agg = df.groupby(keys, dropna=False)[[c for c in PERF_NUM if c in df.columns]].sum()
    out = agg.join(nm.rename("기획전명")).reset_index()
    return out


def load_da_perf_bytes(file_bytes, media="비즈보드", sheet="DA로우_RAW"):
    """대용량 DA로우 엑셀 bytes → (기획전번호, 소재N) 소재별 성과. openpyxl 스트리밍 집계
    (45MB 전체 누적도 pandas 전체 로드 없이 메모리 효율적으로 처리)."""
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = [("" if h is None else str(h).strip()) for h in next(it)]
    idx = {}
    for i, h in enumerate(header):
        idx.setdefault(h, i)
    af_i, pid_i = idx.get("구분_AF코드이름"), idx.get("구분_기획전 번호")
    br_i = idx.get("구분_브랜드/기획전")
    met_i = {c: idx[c] for c in PERF_NUM if c in idx}
    tok = MEDIA_TOKENS.get(media)
    agg, names = {}, {}
    for row in it:
        name = row[af_i] if af_i is not None and af_i < len(row) else None
        if not name:
            continue
        name = str(name)
        if tok and tok not in name:
            continue
        m = re.search(r"_(\d+)\s*$", name)
        if not m:
            continue
        N = int(m.group(1))
        pv = row[pid_i] if pid_i is not None and pid_i < len(row) else None
        try:
            pid = int(float(pv))
        except (TypeError, ValueError):
            continue
        key = (pid, N)
        d = agg.setdefault(key, {c: 0.0 for c in met_i})
        for c, i in met_i.items():
            v = row[i] if i < len(row) else None
            try:
                d[c] += float(v)
            except (TypeError, ValueError):
                pass
        if key not in names and br_i is not None and br_i < len(row) and row[br_i]:
            names[key] = str(row[br_i])
    wb.close()
    rows = [{"기획전번호": k[0], "소재N": k[1], "기획전명": names.get(k, ""), **v}
            for k, v in agg.items()]
    df = pd.DataFrame(rows)
    if not df.empty:
        df["기획전번호"] = df["기획전번호"].astype("Int64")
        df["소재N"] = df["소재N"].astype("Int64")
    return df


def _rate(n, d):
    d = pd.to_numeric(d, errors="coerce")
    return np.where(d > 0, pd.to_numeric(n, errors="coerce") / d, np.nan)


def add_perf_kpis(perf):
    """소재별 성과에 CTR/ROAS/RPS/CR 파생 추가."""
    p = perf.copy()
    p["CTR"]  = _rate(p["지표_클릭수"], p["지표_노출수"])
    p["ROAS"] = _rate(p["지표_순결제거래액"], p["지표_광고비"])
    p["RPS"]  = _rate(p["지표_순결제거래액"], p["지표_노출수"])
    p["CR"]   = _rate(p["지표_순결제고객수"], p["지표_UV(전체)"])
    return p


# ══════════════════════════════════════════════════════════════════════
# 1. 소재안 카피 → 소구형 태깅  (수동 라벨 대신 카피/이미지 기준 재분류)
#    택소노미는 LF 실사용(할인/혜택/인기/공감/상품/브랜드)에 신상/시즌/한정 등을 보강.
# ══════════════════════════════════════════════════════════════════════
# 소구형 사전 — 발송 대시보드 방식(멀티라벨: 매칭되는 소구를 전부 True) 을 DA용으로 세분화.
# 메인+서브 카피를 합친 문자열에 각 정규식을 돌려 소재당 여러 소구를 동시에 태깅한다.
DA_KW = {
    # ── 가격·혜택 축 ──
    "할인율": ["할인", "세일", "오프", "OFF", "off", "반값", "％", "최대 %"],
    "특가·최저가": ["특가", "최저가", "단돈", "균일가", "초특가", "핫딜", "득템", "파격", "역대급"],
    "쿠폰·적립": ["쿠폰", "플러스쿠폰", "적립", "페이백", "캐시백", "포인트", "즉시할인"],
    "사은·증정": ["증정", "사은품", "1+1", "기프트", "덤", "추가증정", "받아가"],
    # ── 희소·긴급 축 ──
    "마감임박": ["오늘", "마지막", "단 7일", "단7일", "단 하루", "단하루", "24시간", "막바지",
               "임박", "까지", "마감", "종료", "곧", "놓치"],
    "한정·단독": ["단독", "한정", "선착순", "리미티드", "독점", "LF몰 단독", "온라인 단독", "조기"],
    # ── 상품·브랜드 축 ──
    "신상·출시": ["신상", "신제품", "출시", "입고", "재입고", "론칭", "컬렉션", "NEW", "new", "New",
               "드롭", "예약판매", "예판", "선공개"],
    "베스트·인기": ["베스트", "BEST", "Best", "인기", "랭킹", "1위", "TOP", "완판", "품절대란",
                "화제", "스테디", "많이 팔린", "사랑받"],
    "브랜드소구": ["브랜드", "공식", "브랜드데이", "브랜드위크", "오리지널"],
    "상품스펙": ["린넨", "냉감", "썸머", "방수", "경량", "코튼", "캐시미어", "기능성", "소재",
              "쿨링", "통기", "신축", "선쉴드", "UV"],
    # ── 감성·상황 축 ──
    "공감·페인포인트": ["그만", "고민", "걱정", "말고", "안 입", "뭐 입", "이제", "후회", "라면",
                   "때", "고민 끝", "지친"],
    "추천·큐레이션": ["추천", "MD", "코디", "스타일링", "엄선", "PICK", "Pick", "픽", "제안",
                  "골라", "완성", "가이드", "셋업"],
    "시즌·시의성": ["여름", "폭염", "무더위", "장마", "바캉스", "휴가", "연휴", "역시즌", "환절기",
                "간절기", "겨울", "가을", "봄", "8월", "신년", "명절", "필드", "라운딩"],
    "개인화·타겟": ["회원님", "고객님", "님,", "VIP", "전용", "멤버십", "여성", "남성", "키즈", "우리아이"],
    "후기·신뢰": ["후기", "리뷰", "재구매", "검증", "평점", "증명", "실착"],
}
KW_RE = {k: re.compile("|".join(re.escape(w) for w in ws)) for k, ws in DA_KW.items()}
PCT_RE    = re.compile(r"\d{1,3}\s*[%％]")            # 할인율(60%, ~80%)
PRICE_RE  = re.compile(r"\d{1,3}(,\d{3})+\s*원|\d+\s*만\s*원|단돈")  # 가격 언급
NUM_RE    = re.compile(r"\d")
ENG_RE    = re.compile(r"[A-Za-z]{2,}")
EMOJI_RE  = re.compile("[\U0001F000-\U0001FAFF\U00002600-\U000027BF\U0001F1E6-\U0001F1FF]")

# 형식(form) 속성 — 카피 '내용'이 아니라 '표현 방식'. 소구형과 별개 축으로 함께 분석.
FORM_TAGS = ["질문형", "느낌표", "숫자강조", "영문포함", "이모지"]
TAG_BOOLS = list(DA_KW.keys()) + FORM_TAGS


def _s(v):
    return "" if v is None or (isinstance(v, float) and pd.isna(v)) else str(v)


def tag_copy(main_copy, sub_copy=""):
    """메인+서브 카피 → 소구형(멀티라벨) + 형식 속성 dict. 매칭되는 소구를 전부 True."""
    t = (_s(main_copy) + " " + _s(sub_copy)).strip()
    d = {k: bool(KW_RE[k].search(t)) for k in DA_KW}
    d["할인율"] = d["할인율"] or bool(PCT_RE.search(t))
    d["특가·최저가"] = d["특가·최저가"] or bool(PRICE_RE.search(t))
    # 형식 속성
    d["질문형"] = ("?" in t or "？" in t)
    d["느낌표"] = ("!" in t or "！" in t)
    d["숫자강조"] = bool(NUM_RE.search(t))
    d["영문포함"] = bool(ENG_RE.search(t))
    d["이모지"] = bool(EMOJI_RE.search(t))
    return d


# 소재당 소구 개수(멀티라벨의 '몇 개를 담았나')·주 소구(가장 앞선 축) 파생
_SOGU_KEYS = list(DA_KW.keys())


def add_tags(sojae_df, main_col="메인카피", sub_col="서브카피"):
    """소재안 DataFrame 에 소구형(멀티라벨)+형식 태그(bool) 컬럼과 소구개수를 추가."""
    df = sojae_df.copy()
    mains = df[main_col] if main_col in df.columns else pd.Series([""] * len(df), index=df.index)
    subs  = df[sub_col] if sub_col in df.columns else pd.Series([""] * len(df), index=df.index)
    tags = pd.DataFrame([tag_copy(m, s) for m, s in zip(mains, subs)], index=df.index)
    out = pd.concat([df, tags], axis=1)
    out["소구개수"] = out[_SOGU_KEYS].sum(axis=1).astype(int)     # 소재당 담은 소구 수
    return out


# ══════════════════════════════════════════════════════════════════════
# 1-B. 이미지 속성 (소재안 컬럼에서 유도 — Vision 없이 지금 가능한 것)
# ══════════════════════════════════════════════════════════════════════
_CODE_RE = re.compile(r"[A-Z]{2,}\d[A-Z0-9]{2,}")     # 상품코드 형태(예: HZTS6F891BK)


def derive_attrs(sojae_df):
    """소재안 컬럼 기반 이미지 속성 유도: 상품수·상품수구간·로고구분.
    (타이포/레이아웃/실측 상품이미지수 등은 Vision 필요 — v2)."""
    df = sojae_df.copy()
    codes = df["상품코드"].map(_s) if "상품코드" in df.columns else pd.Series([""] * len(df))
    df["상품수"] = codes.map(lambda s: len(_CODE_RE.findall(s)))
    df["상품수구간"] = pd.cut(df["상품수"], [-1, 0, 1, 2, 99],
                          labels=["0(로고/카피형)", "1개", "2개", "3개+"]).astype(str)
    if "로고" in df.columns:
        df["로고구분"] = df["로고"].map(lambda v: {"LF": "LF로고", "BR": "브랜드로고"}
                                    .get(_s(v).strip().upper(), _s(v).strip() or "없음"))
    # ── 카피 심화 속성 (메인+서브) ──
    main = df["메인카피"].map(_s) if "메인카피" in df.columns else pd.Series([""] * len(df))
    both = (main + " " + (df["서브카피"].map(_s) if "서브카피" in df.columns else "")).str.strip()
    df["할인율구간"] = both.map(discount_bucket)
    df["문장형"] = main.map(sentence_type)
    df["브랜드노출수"] = both.map(count_brands)
    df["브랜드구성"] = df["브랜드노출수"].map(lambda n: "멀티브랜드" if n >= 2 else ("단일브랜드" if n == 1 else "브랜드무노출"))
    return df


# 카피 심화 분류 헬퍼 ──────────────────────────────────────────
BRANDS = ["헤지스", "닥스", "질스튜어트", "질 스튜어트", "킨", "바네사브루노", "이자벨마랑", "데상트",
          "르꼬끄", "TNGT", "티엔지티", "라코스테", "에트로", "프라다", "질샌더", "지방시", "페라가모",
          "바버", "빈스", "알레그리", "리복", "크래쉬배기지", "벨류엣", "콜롬보", "애타트", "아떼",
          "PXG", "타이틀리스트", "챌린저", "볼빅", "에코", "더블플래그", "리스", "마에스트로", "라움"]
_BRAND_RE = re.compile("|".join(re.escape(b) for b in sorted(BRANDS, key=len, reverse=True)))
_PCT_ALL = re.compile(r"(\d{1,3})\s*[%％]")
# 명령·청유형 어미/표현
_IMPER_RE = re.compile(r"(하세요|보세요|담으세요|가세요|해보세요|만나보|받아가|보러가|확인하|시작|주목|"
                       r"놓치지|드세요|하러|즐기세요|누리세요|잡으세요)")


def discount_bucket(text):
    """카피의 최대 할인율 → 구간. 없으면 '할인율없음'."""
    t = _s(text)
    vals = [int(m) for m in _PCT_ALL.findall(t) if int(m) <= 100]
    if "반값" in t:
        vals.append(50)
    if not vals:
        return "할인율없음"
    mx = max(vals)
    if mx <= 30: return "~30%"
    if mx <= 50: return "~50%"
    if mx <= 70: return "~70%"
    return "~80%+"


def sentence_type(main):
    """메인카피 문장형: 감탄형/의문형/명령·청유형/평서형."""
    t = _s(main).strip()
    if "!" in t or "！" in t: return "감탄형"
    if "?" in t or "？" in t: return "의문형"
    if _IMPER_RE.search(t):   return "명령·청유형"
    return "평서형"


def count_brands(text):
    """카피에 등장한 브랜드명 개수(중복 제거)."""
    return len(set(_BRAND_RE.findall(_s(text))))


# ══════════════════════════════════════════════════════════════════════
# 2. 조인: 소재안(카피·소구형·이미지속성) ↔ 실적(성과)
# ══════════════════════════════════════════════════════════════════════
def join_sojae_perf(sojae_df, perf_df, pid_col="기획전번호", order_col="소재N"):
    """(기획전번호, 소재N) 기준 조인. 소재안=왼쪽(카피/속성), 실적=오른쪽(성과)."""
    s = sojae_df.copy()
    s[pid_col] = pd.to_numeric(s[pid_col], errors="coerce").astype("Int64")
    s[order_col] = pd.to_numeric(s[order_col], errors="coerce").astype("Int64")
    return s.merge(perf_df, on=[pid_col, order_col], how="inner", suffixes=("", "_실적"))


# ══════════════════════════════════════════════════════════════════════
# 3. 소구형/이미지속성별 성과 + 유의성  (발송 대시보드 통계층 이식)
# ══════════════════════════════════════════════════════════════════════
def _welch_p(a, b):
    a = np.asarray(a, float); a = a[~np.isnan(a)]
    b = np.asarray(b, float); b = b[~np.isnan(b)]
    if len(a) < 3 or len(b) < 3:
        return None
    try:
        from scipy import stats
        return float(stats.ttest_ind(a, b, equal_var=False).pvalue)
    except Exception:
        return None


def _cohen_d(a, b):
    a = np.asarray(a, float); a = a[~np.isnan(a)]
    b = np.asarray(b, float); b = b[~np.isnan(b)]
    if len(a) < 2 or len(b) < 2:
        return np.nan
    sp = np.sqrt(((len(a)-1)*a.var(ddof=1) + (len(b)-1)*b.var(ddof=1)) / (len(a)+len(b)-2))
    return (a.mean() - b.mean()) / sp if sp else np.nan


def fdr_bh(pvals):
    idx = [i for i, p in enumerate(pvals) if p is not None]
    m = len(idx); out = [None]*len(pvals)
    if not m:
        return out
    prev = 1.0
    for rank, i in enumerate(sorted(idx, key=lambda i: pvals[i], reverse=True), 1):
        prev = min(prev, pvals[i] * m / (m - rank + 1)); out[i] = prev
    return out


def appeal_perf(joined, metric_label, tags=None, min_n=5):
    """소구형(또는 임의 bool 태그)별 '보유 vs 미보유' 가중성과 + 유의성."""
    num, den, higher = METRICS[metric_label]
    tags = tags or [t for t in TAG_BOOLS if t in joined.columns]
    per_row = pd.Series(_rate(joined[num], joined[den]), index=joined.index)
    rows, pv = [], []
    for tag in tags:
        has = joined[tag].fillna(False).astype(bool)
        nh, nn = int(has.sum()), int((~has).sum())
        if nh < min_n or nn < min_n:
            continue
        wh = _rate(joined.loc[has, num].sum(), joined.loc[has, den].sum())
        wn = _rate(joined.loc[~has, num].sum(), joined.loc[~has, den].sum())
        p = _welch_p(per_row[has].values, per_row[~has].values)
        rows.append({"태그": tag, "보유가중": float(wh), "보유n": nh, "미보유가중": float(wn),
                     "미보유n": nn, "차이": float(wh)-float(wn),
                     "효과크기": _cohen_d(per_row[has].values, per_row[~has].values), "p": p})
        pv.append(p)
    for r, q in zip(rows, fdr_bh(pv)):
        r["q"] = q
    out = pd.DataFrame(rows)
    return out.sort_values("차이", ascending=not higher).reset_index(drop=True) if len(out) else out


def cat_perf(joined, cat_col, metric_label, min_n=3):
    """범주형 컬럼(이미지유형·브랜드·소구형_수동 등) 값별 가중 성과."""
    num, den, higher = METRICS[metric_label]
    rows = []
    for val, sub in joined.groupby(cat_col):
        if not str(val).strip() or len(sub) < min_n:
            continue
        rows.append({cat_col: str(val), "n": len(sub),
                     metric_label: float(_rate(sub[num].sum(), sub[den].sum())),
                     "노출": float(pd.to_numeric(sub["지표_노출수"], errors="coerce").sum())})
    out = pd.DataFrame(rows)
    return out.sort_values(metric_label, ascending=not higher).reset_index(drop=True) if len(out) else out


def sig_label(q):
    if q is None or (isinstance(q, float) and pd.isna(q)):
        return "표본부족"
    if q < 0.01: return "p<0.01 · 신뢰가능"
    if q < 0.05: return "p<0.05 · 유의함"
    if q < 0.10: return "p<0.10 · 약한신호"
    return f"q={q:.2f} · 유의하지않음"


# ══════════════════════════════════════════════════════════════════════
# 5. 인사이트 요약 → Gemini 카피 제안
# ══════════════════════════════════════════════════════════════════════
def build_insight_summary(joined, metric_label, min_n=4):
    """조인 데이터에서 '이기는 패턴'을 요약 텍스트로. (카피 제안 프롬프트용)"""
    higher = METRICS[metric_label][2]
    lines = [f"[성과 지표: {metric_label} · 분석 소재 {len(joined)}건]"]

    ap = appeal_perf(joined, metric_label, tags=list(DA_KW), min_n=min_n)
    if len(ap):
        win = ap[(ap["차이"] > 0) == higher].head(4)["태그"].tolist()
        lose = ap[(ap["차이"] > 0) != higher].tail(3)["태그"].tolist()
        if win:  lines.append("· 잘 되는 소구형: " + ", ".join(win))
        if lose: lines.append("· 약한 소구형: " + ", ".join(lose))

    kp = keyword_perf(joined, metric_label, min_n=max(5, min_n), top=10)
    if len(kp):
        goodw = kp[(kp["차이"] > 0) == higher]["단어"].head(8).tolist()
        if goodw: lines.append("· 성과 좋은 단어: " + ", ".join(goodw))

    for col, lab in [("이미지유형", "이미지유형"), ("문장형", "카피 문장형"),
                     ("할인율구간", "할인율"), ("브랜드구성", "브랜드 노출")]:
        if col in joined.columns:
            cp = cat_perf(joined, col, metric_label, min_n=min_n)
            if len(cp):
                lines.append(f"· 잘 되는 {lab}: {cp.iloc[0][col]}")

    cx = cross_perf(joined, list(DA_KW), "이미지유형", metric_label, min_n=3)
    if not cx.empty:
        stacked = cx.stack()
        if len(stacked):
            best = stacked.idxmax() if higher else stacked.idxmin()
            lines.append(f"· 최고 조합: 소구 '{best[0]}' × 이미지 '{best[1]}'")
    return "\n".join(lines)


def gemini_text(prompt, api_key=None, model="gemini-3.6-flash", tries=4):
    """텍스트 생성 (카피 제안용). 429 백오프. google-genai 필요."""
    import time
    key = api_key
    if not key:
        try:
            import streamlit as st
            key = st.secrets.get("gemini_api_key")
        except Exception:
            key = None
    if not key:
        return None, "Gemini 키가 없습니다 (secrets: gemini_api_key)."
    try:
        from google import genai
    except Exception as e:
        return None, f"google-genai 미설치: {e}"
    client = genai.Client(api_key=key)
    for attempt in range(tries):
        try:
            r = client.models.generate_content(model=model, contents=[prompt])
            return (r.text or ""), None
        except Exception as e:
            msg = str(e)
            if "429" in msg or "RESOURCE_EXHAUSTED" in msg:
                time.sleep(min(2 ** attempt * 5, 40)); continue
            return None, msg[:180]
    return None, "429 재시도 초과"


def suggest_copy(joined, metric_label, context="", n=5, api_key=None, model="gemini-3.6-flash"):
    """인사이트 기반 신규 배너 카피 n개 제안. 반환: (list[dict], error)."""
    import json as _json
    summary = build_insight_summary(joined, metric_label)
    prompt = (
        "너는 LF몰 디스플레이(비즈보드) 광고의 시니어 카피라이터다.\n"
        "아래는 최근 실제 성과 데이터에서 뽑은 '이기는 패턴' 요약이다:\n\n"
        f"{summary}\n\n"
        f"[추가 조건] {context or '특정 브랜드/기획전 지정 없음 — 범용 제안'}\n\n"
        f"이 인사이트(잘 되는 소구형·단어·이미지패턴)를 적극 반영해 신규 비즈보드 배너 카피 {n}개를 제안하라.\n"
        "규칙: 메인카피는 14자 이내, 서브카피는 18자 이내(한글 기준), 실제 광고 문구처럼 자연스럽게.\n"
        "JSON 배열만 출력(설명 금지):\n"
        '[{"메인카피":"..","서브카피":"..","활용패턴":"반영한 소구형/단어/이미지","기대효과":"왜 잘 될지 한줄"}]'
    )
    txt, err = gemini_text(prompt, api_key=api_key, model=model)
    if err:
        return [], err
    m = re.search(r"\[.*\]", txt, re.S)
    if not m:
        return [], "카피 JSON 파싱 실패"
    try:
        return _json.loads(m.group(0)), None
    except Exception as e:
        return [], f"JSON 파싱 오류: {e}"


# ── 단어 단위 성과 (발송 대시보드 keyword_perf 이식) ──────────────
_STOP = set(["외", "및", "the", "of", "up", "to", "for", "with", "and", "LF몰", "LFmall",
             "지금", "더", "이", "그", "저", "수", "것", "때", "안"])
_WORD_RE = re.compile(r"[가-힣A-Za-z][가-힣A-Za-z0-9]{1,}")


def keyword_perf(joined, metric_label, main_col="메인카피", sub_col="서브카피",
                 min_n=5, top=30):
    """카피 단어별 '포함 vs 미포함' 가중 성과 + 유의성. 어떤 단어가 지표를 올리나."""
    num, den, higher = METRICS[metric_label]
    texts = (joined[main_col].map(_s) + " " + joined.get(sub_col, "").map(_s)) \
        if main_col in joined.columns else pd.Series([""] * len(joined), index=joined.index)
    word_sets = texts.map(lambda t: set(w for w in _WORD_RE.findall(t)
                                        if w not in _STOP and len(w) >= 2))
    from collections import Counter
    cnt = Counter(w for ws in word_sets for w in ws)
    per_row = pd.Series(_rate(joined[num], joined[den]), index=joined.index)
    rows, pv = [], []
    for w, c in cnt.items():
        if c < min_n or len(joined) - c < 3:
            continue
        has = word_sets.map(lambda s: w in s)
        wh = _rate(joined.loc[has, num].sum(), joined.loc[has, den].sum())
        wn = _rate(joined.loc[~has, num].sum(), joined.loc[~has, den].sum())
        p = _welch_p(per_row[has].values, per_row[~has].values)
        rows.append({"단어": w, "포함가중": float(wh), "포함n": int(c),
                     "미포함가중": float(wn), "차이": float(wh) - float(wn), "p": p})
        pv.append(p)
    for r, q in zip(rows, fdr_bh(pv)):
        r["q"] = q
    out = pd.DataFrame(rows)
    if out.empty:
        return out
    return out.sort_values("차이", ascending=not higher).head(top).reset_index(drop=True)


# ── 교차분석: 태그(bool) × 범주형 → 가중 성과 매트릭스 ──────────────
def cross_perf(joined, tag_cols, cat_col, metric_label, min_n=3):
    """행=소구형(각 bool True인 소재), 열=범주형 값. 셀=가중 성과(표본 min_n 미만은 NaN)."""
    num, den, _ = METRICS[metric_label]
    cats = [c for c in joined[cat_col].dropna().unique() if str(c).strip()]
    out = {}
    for tag in tag_cols:
        if tag not in joined.columns:
            continue
        sub_t = joined[joined[tag].fillna(False).astype(bool)]
        row = {}
        for cv in cats:
            cell = sub_t[sub_t[cat_col] == cv]
            row[cv] = float(_rate(cell[num].sum(), cell[den].sum())) if len(cell) >= min_n else np.nan
        out[tag] = row
    return pd.DataFrame(out).T  # index=소구형, columns=범주


# ── OLS 순효과: 교란(브랜드·카테고리·이미지유형) 통제 후 소구형 계수 ──
def ols_effects(joined, tag_cols, control_cols, metric_label, min_n=4):
    """소구형 더미 + 통제변수 더미로 metric 회귀 → 소구형별 순효과(계수)와 근사 유의성.
    통제 후에도 남는 소구형의 순수 기여를 본다(단순 비교의 착시 제거)."""
    num, den, higher = METRICS[metric_label]
    y = pd.Series(_rate(joined[num], joined[den]), index=joined.index)
    mask = y.notna()
    d = joined[mask].copy(); y = y[mask].values
    if len(y) < min_n * 2:
        return pd.DataFrame()
    X_parts, names = [np.ones((len(d), 1))], ["const"]
    tags_used = []
    for t in tag_cols:
        if t in d.columns and d[t].fillna(False).astype(bool).sum() >= min_n:
            X_parts.append(d[t].fillna(False).astype(float).values.reshape(-1, 1))
            names.append(t); tags_used.append(t)
    for c in control_cols:
        if c not in d.columns:
            continue
        dummies = pd.get_dummies(d[c].astype(str), prefix=c, drop_first=True)
        if len(dummies.columns):
            X_parts.append(dummies.values.astype(float)); names += list(dummies.columns)
    X = np.hstack(X_parts)
    try:
        beta, _, _, _ = np.linalg.lstsq(X, y, rcond=None)
        resid = y - X @ beta
        dofr = max(len(y) - X.shape[1], 1)
        sigma2 = (resid @ resid) / dofr
        XtX_inv = np.linalg.pinv(X.T @ X)
        se = np.sqrt(np.maximum(np.diag(XtX_inv) * sigma2, 0))
    except Exception:
        return pd.DataFrame()
    rows = []
    for t in tags_used:
        i = names.index(t)
        coef, s = beta[i], se[i]
        from math import erf, sqrt
        z = coef / s if s > 0 else 0.0
        p = 2 * (1 - 0.5 * (1 + erf(abs(z) / sqrt(2))))
        rows.append({"소구형": t, "순효과(계수)": float(coef), "p": float(p)})
    out = pd.DataFrame(rows)
    return out.sort_values("순효과(계수)", ascending=not higher).reset_index(drop=True) if len(out) else out


# ══════════════════════════════════════════════════════════════════════
# 4. Streamlit UI (개요) — 조인된 소재×성과 위에서 렌더
# ══════════════════════════════════════════════════════════════════════
def render(sojae_df, raw_perf_df, media="비즈보드"):
    import streamlit as st
    st.title("📝 DA 카피 분석·제안")
    if sojae_df is None or raw_perf_df is None:
        st.info("소재안(카피)과 실적(DA로우)을 모두 올리면 소재별로 조인해 분석해요.")
        return
    perf = add_perf_kpis(load_da_perf(raw_perf_df, media=media))
    tagged = add_tags(sojae_df)
    joined = join_sojae_perf(tagged, perf)
    st.caption(f"조인된 소재 {len(joined):,}건 · 매체: {media}")
    metric = st.selectbox("성과 주지표", list(METRICS), index=0)
    ap = appeal_perf(joined, metric)
    if ap.empty:
        st.warning("표본이 부족해요. 기간을 넓히거나 매체를 바꿔보세요."); return
    st.dataframe(ap, use_container_width=True, hide_index=True)


if __name__ == "__main__":
    import sys
    # 실적 로더 스모크 테스트: 실제 DA로우 파일 경로를 인자로 주면 소재별 성과를 출력
    if len(sys.argv) > 1:
        raw = pd.read_excel(sys.argv[1], sheet_name="DA로우_RAW", header=0, engine="openpyxl")
        perf = add_perf_kpis(load_da_perf(raw, media="비즈보드"))
        print("소재행:", len(perf), "· 기획전:", perf["기획전번호"].nunique())
        top = perf[perf["지표_노출수"] > 3000].copy()
        top["CTR%"] = (top["CTR"]*100).round(2); top["ROAS%"] = (top["ROAS"]*100).round(0)
        print(top.sort_values("지표_노출수", ascending=False)
              [["기획전번호", "기획전명", "소재N", "지표_노출수", "지표_클릭수", "CTR%", "ROAS%"]].head(12).to_string(index=False))
    else:
        # 태깅 스모크 테스트
        print("tags:", tag_copy("데상트&르꼬끄 여름골프 ~60%", "단독특가 시원한 라운딩!"))
